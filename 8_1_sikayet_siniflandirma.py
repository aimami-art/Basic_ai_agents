"""
kullanicidan gelen şikayetlerin analiz edilmesi, kategorize edilmesi, 
ilgili birime yönlendirilmesi ve uygun yanıt verilmesi sistemi gerceklestirelim.

+ otomatik mail (simulasyon) + raporlama (.xlsx) 

"""

from langchain.chat_models import ChatOpenAI
from langchain.prompts import PromptTemplate
from langchain.output_parsers import StructuredOutputParser, ResponseSchema
from dotenv import load_dotenv
import os
import warnings
from openpyxl import Workbook, load_workbook
from datetime import datetime

warnings.filterwarnings("ignore")

# .env dosyasını yükle
load_dotenv()
# OpenAI API anahtarını al
api_key = os.getenv("OPENAI_API_KEY")

# --- 1. LLM Modeli ---
llm = ChatOpenAI(
    model_name="gpt-3.5-turbo",
    openai_api_key=api_key,
    temperature=0.2 # Düşük sıcaklık, daha kararlı ve tutarlı cevaplar için
)

# --- 2. Beklenen JSON formatı için şema tanımları ---
response_schemas = [
    ResponseSchema(name="kategori", description="Şikayet hangi kategoriye ait? (kargo, iade, ödeme, teknik, iletişim)"),
    ResponseSchema(name="yonlendirme", description="Hangi departmana yönlendirilmeli?"),
    ResponseSchema(name="cevap", description="Kullanıcıya uygun tathmin edici bir yanıt metni")
]

# Parser
parser = StructuredOutputParser.from_response_schemas(response_schemas)
format_instructions = parser.get_format_instructions()

# --- 3. Prompt Template ---
template = PromptTemplate(
    input_variables=["sikayet"],
    partial_variables={"format_instructions": format_instructions},
    template="""
            Bir kullanıcı şikayeti aldın. 
            Bu şikayeti anlamlandır, uygun kategoriyi belirle, 
            yönlendirme yap ve kullanıcıya kısa bir açıklama yaz.

            {format_instructions}

            Şikayet: {sikayet}
            """
)

# --- 4. Kullanıcı girdisi (örnek) ---
sikayet = input("Şikayetinizi yazınız: ")

# Prompt'u oluştur
prompt = template.format(sikayet=sikayet)

# LLM'den yanıt al
response = llm.invoke(prompt)

# Yapılandırılmış veriyi ayrıştır
result = parser.parse(response.content)

# excel e kaydetme
excel_dosya = "sikayet_raporu.xlsx"
if not os.path.exists(excel_dosya):
    wb = Workbook()
    ws = wb.active
    ws.title = "Şikayet Raporu"
    ws.append(["Zaman", "Şikayet","Kategori", "Yönlendirme", "Cevap"])  # Başlık satırı
else:
    wb = load_workbook(excel_dosya)
    ws = wb.active

ws.append([
    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),  # Zaman damgası
    sikayet,  # Kullanıcının şikayeti
    result["kategori"],  # Kategori
    result["yonlendirme"],  # Yönlendirme
    result["cevap"]  # Cevap
])
wb.save(excel_dosya)  # Dosyayı kaydet


# Sonucu göster
print("Otomatik e-posta cevabı")
print(f"Kime: kullanici@example.com")
print("Konu: Şikayetiniz Hakkında")
print("İçerik:")
print(result["cevap"])